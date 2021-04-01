#!/usr/bin/env python
# -*- coding: UTF-8 -*-

import re
import os
from collections import defaultdict
from openpyxl import Workbook


class SaSinglePassword:
    def __init__(self, pattern):
        self.password = ''
        self.pattern = pattern          # 格式 ：分析特殊字符还是大写字母
        self.otherStr = []              # 排除的口令，存储password
        self.findStr = []               # 从口令中提取出的子串列表
        self.total = 0                  # 文件中的口令总数量
        self.count = 0                  # 代表count之和，符合条件的口令总数量
        self.count_no = 0               # 代表不含（特殊或大写）字符的
        self.count_alpha = 0            # 单个字符总数量
        self.posDic = defaultdict(int)  # 存储位置数量  [{position:count},... ] begin,media,end代表0，1，2
        self.findDic = defaultdict(int)     # 子串长度字典 {length:count},...
        self.findSDic = defaultdict(int)    # 单个字符的统计  [ {'*' : 78},... ]
        self.min_length = 999           # 口令中的最短子串长度
        self.max_length = 0             # 口令中的最长子串长度

    def set_find_str(self):         # 匹配字符串中的对应模式，返回列表
        self.findStr = re.findall(self.pattern, self.password)

    def set_dic(self, password):
        self.password = password
        self.total += 1
        self.set_find_str()         # 提取子串
        sub = self.findStr
        if len(sub) == 1:
            self.count += 1
            # 判断位置
            pos = self.get_position(sub[0])
            self.posDic[pos] += 1
            # 判断长度
            sub_len = len(sub[0])
            self.findDic[sub_len] += 1
            self.min_length = min(sub_len, self.min_length)
            self.max_length = max(sub_len, self.max_length)
            # 判断单个字母
            for a in sub[0]:
                self.count_alpha += 1
                self.findSDic[a] += 1
        elif len(sub) == 0:
            self.count_no += 1
        else:
            self.otherStr.append(password)

    def get_position(self, sub):
        i = self.password.index(sub)
        if i == 0:
            return 'Begin'
        elif i+len(sub)-1 == len(self.password):
            return 'End'
        else:
            return 'Mediate'


class Compute:
    def __init__(self, in_file, pattern):
        self.dir = ''
        self.pattern = pattern
        self.sas = SaSinglePassword(pattern)
        self.encoding = 'UTF-8'         # 设置文件编码
        self.in_file = in_file
        self.label = self.in_file.split('/')[-1].split('.')[0]
        self.out_txt = 'other.txt'
        self.decimal = 4                # 默认保存到小数点后4位
        self.topN = 10
        self.out_excel_pos = 'position.xls'
        self.out_excel_length = 'length.xls'
        self.out_excel_single = 'single.xls'

    def read_file(self):
        print('正在读取文件{0}...'.format(self.in_file))
        with open(self.in_file, 'r', encoding=self.encoding) as f:
            for line in f:
                line = line.strip('\n')
                yield line

    def txt_write_list(self, result):  # 写不合法口令到文件
        with open(os.path.join(self.dir, self.out_txt), 'w', encoding=self.encoding) as f:
            for r in result:
                f.write(r + '\n')
        print('文件{0}写入完毕!'.format(self.out_txt))

    def excel_write_list(self, result, count, file):
        wb = Workbook()
        sheet = wb.active
        sheet.cell(1, 1, '分析文件：{0}'.format(self.in_file))
        left = self.sas.total-self.sas.count_no-self.sas.count
        sheet.cell(2, 1, '总口令数：{0}，含一种位置的口令数：{1}，含多个位置的口令数：{2}，不含（特殊或大写）的口令数：{3}\n'.format(self.sas.total, self.sas.count, left, self.sas.count_no))
        sheet.cell(3, 1, '排名Top {0}\n'.format(self.topN))
        row_index = 4
        for e in result:
            e_3 = round(100 * e[1] / count, self.decimal)
            sheet.cell(row_index, 1, e[0])
            sheet.cell(row_index, 2, e[1])
            sheet.cell(row_index, 3, '{0}%'.format(e_3))
            row_index += 1
        wb.save(os.path.join(self.dir, file))
        print('文件{0}数据写入完毕!'.format(file))

    def run(self):
        kind = 'Special' if self.pattern == '(\W+)' else 'Upper'
        self.dir = self.label + kind
        print(self.dir)
        if not os.path.exists(self.dir):
            os.makedirs(self.dir)
            print("文件夹{0}创建成功！".format(self.dir))

        for line in self.read_file():
            self.sas.set_dic(line)

        # 第一个文件：不合法口令文件 #
        self.txt_write_list(self.sas.otherStr)

        # 第二个文件：合法口令 位置占比文件 #
        res_list = self.get_sorted_list(self.sas.posDic)
        self.excel_write_list(res_list, self.sas.count, self.out_excel_pos)

        # 第三个文件：合法口令 长度占比文件 #
        res_list = self.get_sorted_list(self.sas.findDic)
        self.excel_write_list(res_list, self.sas.count, self.out_excel_length)

        # 第三个文件：合法口令 长度占比文件 #
        res_list = self.get_sorted_list(self.sas.findSDic)
        self.excel_write_list(res_list, self.sas.count_alpha, self.out_excel_single)

    def get_sorted_list(self, dic):  # 字典按值排序，返回排序后的前n个值
        a = dic
        n = self.topN
        result_list = sorted(a.items(), key=lambda item: item[1], reverse=True)         # 降序排序
        return result_list[:n]


# 生成两个文件夹 Special和Upper
if __name__ == '__main__':
    # file = 'Hack/Tuscl.txt'
    dir = 'Hack'
    files = os.listdir(dir)  # 获得文件夹中所有文件的名称列表
    for f in files:
        path = os.path.join(dir, f)
        Compute(path, '(\W+)').run()        # 统计特殊字符
        Compute(path, '([A-Z]+)').run()     # 统计大写字母
