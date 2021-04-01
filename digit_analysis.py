#!/usr/bin/env python
# -*- coding: UTF-8 -*-

import os
import re
from openpyxl import Workbook


class DaSinglePwd:  # 密码数字串分析类
    def __init__(self):
        self.password = ''
        self.digitStr = []
        self.total = 0  # 所有的口令数量
        self.count = 0  # 已经分析出的合法的密码数量
        self.discontinuous = []  # 存放不连续的口令
        self.min_dig_length = 999  # 口令中的最短数字串长度
        self.max_dig_length = 0  # 口令中的最长数字串长度
        self.decimal = 4  # 默认保存到小数点后4位
        self.lengthDic = {}  # 数字串长度字典   key:length value:[数字段总数量,{数字段:个数}]
        self.lengthDicPer = {}  # 数字串长度字典百分比

    def set_digit_str(self):
        self.digitStr = re.findall('(\d+)', self.password)  # 匹配字符串中的数字，返回列表

    def set_digit_dic(self, password):  # ***主要引用函数***
        self.password = password
        self.total += 1
        self.set_digit_str()  # 提取数字串
        d = self.digitStr
        if len(d) == 1:  # 只有一个数字串时，算合法,才加入字典集合
            self.count += 1
            digit_len = len(d[0])
            self.min_dig_length = min(digit_len, self.min_dig_length)
            self.max_dig_length = max(digit_len, self.max_dig_length)
            if digit_len in self.lengthDic.keys():
                self.lengthDic[digit_len][0] += 1  # 长度数字串个数加1
                if d[0] in self.lengthDic[digit_len][1].keys():
                    self.lengthDic[digit_len][1][d[0]] += 1  # 对应数字段个数加1
                else:
                    self.lengthDic[digit_len][1][d[0]] = 1
            else:
                self.lengthDic[digit_len] = [1, {d[0]: 1}]
        elif len(d) > 1:
            self.discontinuous.append(password)

    def set_digit_dic_per(self):
        import copy
        self.lengthDicPer = copy.deepcopy(self.lengthDic)  # 注意：深复制
        for length in range(self.min_dig_length, self.max_dig_length + 1):
            if length in self.lengthDic.keys():
                dic = self.lengthDic[length][1]
                for key in dic:
                    self.lengthDicPer[length][1][key] = round(
                        100 * self.lengthDic[length][1][key] / self.lengthDic[length][0], self.decimal)

    def get_digit_dic(self):  # 返回字典
        return self.lengthDic

    def get_digit_dic_per(self):  # 返回字典百分比
        return self.lengthDicPer


class RunDa:  # 运行类
    def __init__(self, in_file):
        self.dir = 'xxxDigit'
        self.das = DaSinglePwd()
        self.encoding = 'UTF-8'  # 设置文件编码
        self.in_file = in_file
        self.out_file = 'result.txt'
        self.discontinue_file = 'discontinue.txt'
        self.topN = 10
        self.out_excel = 'result.xlsx'

    def read_file(self):
        print('正在读取文件{0}...'.format(self.in_file))
        with open(self.in_file, 'r', encoding=self.encoding) as f:
            for line in f:
                line = line.strip('\n')
                yield line

    def txt_write_list(self, result):  # 写不连续口令到文件
        with open(os.path.join(self.dir, self.discontinue_file), 'w', encoding=self.encoding) as f:
            for r in result:
                f.write(r + '\n')
        print('文件{0}写入完毕!'.format(self.discontinue_file))

    def txt_write_length(self, result):  # 将列表结果写入文件
        sp = ' ' * 3 + '#' * 3 + ' ' * 3
        with open(os.path.join(self.dir, self.out_file), 'w', encoding=self.encoding) as f:
            f.write('分析文件：{0}\n'.format(self.in_file))
            f.write('总口令数量：{0}，只包含单个数字串的口令数量：{1}\n'.format(self.das.total, self.das.count))
            f.write('排名Top {0}的长度\n'.format(self.topN))
            for e in result:
                e_4 = round(100 * self.das.lengthDicPer[e[0]][0] / self.das.count, self.das.decimal)
                f.write('{0}{1}{2}({3}%)\n'.format(e[0], sp, e[1], e_4))
            print('文件{0}写入完毕'.format(self.out_file))

    def excel_write_length(self, result):  # 将列表结果写入Excel表格
        wb = Workbook()
        sheet = wb.active
        sheet.cell(1, 1, '分析文件：{0}'.format(self.in_file))
        sheet.cell(2, 1, '总口令数量：{0}，只包含单个数字串的口令数量：{1}\n'.format(self.das.total, self.das.count))
        sheet.cell(3, 1, '排名Top {0}的长度\n'.format(self.topN))
        row_index = 4
        for e in result:
            e_4 = round(100 * self.das.lengthDicPer[e[0]][0] / self.das.count, self.das.decimal)
            sheet.cell(row_index, 1, e[0])
            sheet.cell(row_index, 2, e[1])
            sheet.cell(row_index, 3, '{0}%'.format(e_4))
            row_index += 1
        wb.save(os.path.join(self.dir, self.out_excel))
        print('{0}数据写入完毕!'.format(self.out_excel))

    def txt_write_digit(self, length, result):
        sp = ' ' * 3 + '#' * 3 + ' ' * 3
        with open(os.path.join(self.dir, self.out_file), 'w', encoding=self.encoding) as f:
            f.write('数字串长度：{0}，总数量：{1}\n'.format(length, self.das.lengthDic[length][0]))
            f.write('排名Top {0}的数字串\n'.format(self.topN))
            for e in result:
                f.write('{0}{1}{2}({3}%)\n'.format(e[0], sp, e[1], self.das.lengthDicPer[length][1][e[0]]))
        print('文件{0}写入完毕'.format(self.out_file))

    def excel_write_digit(self, length, result, row, col, sheet):
        sheet.cell(row, col, '数字串长度：{0}，总数量：{1}({2}%)\n'.format(length, self.das.lengthDic[length][0], round(100*self.das.lengthDic[length][0]/self.das.count, self.das.decimal)))
        sheet.cell(row + 1, col, '排名Top {0}的数字串：\n'.format(self.topN))
        row_index = row + 2
        for e in result:
            sheet.cell(row_index, col, e[0])
            sheet.cell(row_index, col + 1, e[1])
            sheet.cell(row_index, col + 2, '{0}%'.format(self.das.lengthDicPer[length][1][e[0]]))
            row_index += 1
        print('长度为{0}数字串写入完毕!'.format(length))

    def get_sorted_list(self, dic):  # 字典按值排序，返回排序后的前n个值
        a = dic
        n = self.topN
        result_list = sorted(a.items(), key=lambda item: item[1], reverse=True)  # 降序排序
        return result_list[:n]

    @staticmethod
    def get_dir_files(dirs):  # 获取当前目录下的所有文件
        files = os.listdir(dirs)  # 获得文件夹中所有文件的名称列表
        files_path = []  # 存放path路径中的文件路径
        files_name = []  # 存放文件名字标签 Example:test.txt ===>test
        for f in files:
            path = os.path.join(dirs, f)
            if not os.path.isdir(path):  # 判断是否是文件夹
                files_path.append(path)
                files_name.append(str(f).split('.')[0])
        return files_path

    def run(self):
        self.dir = self.in_file.split('\\')[-1].split('.')[0] + 'Digit'
        if not os.path.exists(self.dir):
            os.makedirs(self.dir)
            print("文件夹{0}创建成功！".format(self.dir))

        for line in self.read_file():
            self.das.set_digit_dic(line)

        # 第一个文件：不连续口令文件 #
        self.txt_write_list(self.das.discontinuous)

        # 第二个文件：连续口令长度占比文件 #
        self.das.set_digit_dic_per()  # 此时已经得到带有百分比的字典了***
        len_range = range(self.das.min_dig_length, self.das.max_dig_length + 1)
        length_dict = {}
        for length in len_range:
            if length in self.das.lengthDic.keys():
                length_dict[length] = self.das.lengthDic[length][0]
        # self.out_file = 'length.txt'
        # self.txt_write_length(res_list)
        self.out_excel = 'length.xlsx'
        res_list = self.get_sorted_list(length_dict)
        self.excel_write_length(res_list)

        # 第三个文件：各个长度的占比 #
        self.out_excel = 'digit.xlsx'
        wb = Workbook()
        sheet = wb.active
        count = 0

        new_dict = {}
        for length in len_range:
            if length in self.das.lengthDic.keys():
                new_dict[length] = self.das.lengthDic[length][0]
        a = new_dict
        result_list = sorted(a.items(), key=lambda item: item[1], reverse=True)  # 降序排序
        res = result_list
        for r in res:
            length = r[0]
            self_dict = self.das.lengthDic[length][1]
            res_list = self.get_sorted_list(self_dict)  # 排序
            row = (count // 5) * (3 + self.topN) + 1  # 计算当前左上角的行值
            col = (count % 5) * 4 + 1  # 计算当前左上角的列值
            self.excel_write_digit(length, res_list, row, col, sheet)
            count += 1
        # for length in len_range:
        #     if length in self.das.lengthDic.keys():
        #         self_dict = self.das.lengthDic[length][1]
        #         res_list = self.get_sorted_list(self_dict)  # 排序
        #         row = (count // 5) * (3 + self.topN) + 1  # 计算当前左上角的行值
        #         col = (count % 5) * 4 + 1  # 计算当前左上角的列值
        #         self.excel_write_digit(length, res_list, row, col, sheet)
        #         count += 1
        wb.save(os.path.join(self.dir, self.out_excel))
        print('{0}数据写入完毕!'.format(self.out_excel))


if __name__ == '__main__':
    # file = 'Hack\\7k7k.txt'
    # R = RunDa(file)
    # R.run()
    # dir = '/home/yh/Passwords/CN'
    dir = 'Hack'
    files = os.listdir(dir)  # 获得文件夹中所有文件的名称列表
    for f in files:
        path = os.path.join(dir, f)
        R = RunDa(path)
        R.run()
